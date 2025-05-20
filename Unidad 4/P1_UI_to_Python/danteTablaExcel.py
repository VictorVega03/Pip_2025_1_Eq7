import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

# Crear un nuevo libro de Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Algoritmo de Lamport"

# Estilos
header_font = Font(bold=True)
border = Border(left=Side(style='thin'),
               right=Side(style='thin'),
               top=Side(style='thin'),
               bottom=Side(style='thin'))
center_aligned = Alignment(horizontal='center')

# Primera tabla - Elección de números
ws['A1'] = "Tabla 1: Elección de números de ticket"
ws['A1'].font = Font(bold=True, size=14)
ws.merge_cells('A1:F1')

# Encabezados
headers = ["Tiempo", "P0", "P1", "P2", "P3", "P4"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.border = border
    cell.alignment = center_aligned

# Datos primera tabla
data1 = [
    ["T0", "ticket[0] = 0, choosing[0] = false", "ticket[1] = 0, choosing[1] = false",
     "ticket[2] = 0, choosing[2] = false", "ticket[3] = 0, choosing[3] = false",
     "ticket[4] = 0, choosing[4] = false"],
    ["T1", "choosing[0] = True", "", "", "", ""],
    ["T2", "Asigna ticket[0] = 1, choosing[0] = False", "", "", "", ""],
    ["T3", "", "choosing[1] = True", "", "choosing[3] = True", ""],
    ["T4", "", "Asigna ticket[1] = 2, choosing[1] = False", "", "Asigna ticket[3] = 2 (mismo que P1), choosing[3] = False", ""],
    ["T5", "", "", "choosing[2] = True", "", ""],
    ["T6", "", "", "Asigna ticket[2] = 3, choosing[2] = false", "", ""],
    ["T7", "", "", "", "", "choosing[4] = True"],
    ["T8", "", "", "", "", "Asigna ticket[4] = 4, choosing[4] = False"],
    ["T9", "--> P0 entra a la sección crítica", "", "", "", ""]
]

for row_idx, row_data in enumerate(data1, 4):
    for col_idx, cell_data in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=cell_data)
        cell.border = border
        if "entra" in str(cell_data):
            cell.font = Font(color="00FF00", bold=True)  # Verde para entradas

# Segunda tabla - Entrada y salida
ws['A14'] = "Tabla 2: Entrada y salida de región crítica"
ws['A14'].font = Font(bold=True, size=14)
ws.merge_cells('A14:F14')

# Datos segunda tabla
data2 = [
    ["T0", "P0 sale, ticket[0] = 0", "", "", "", ""],
    ["T1", "--> P1 entra a la sección crítica", "", "", "", ""],
    ["T2", "P1 sale, ticket[1] = 0", "", "", "", ""],
    ["T3", "--> P3 entra a la sección crítica (porque ya pasó P1)", "", "", "", ""],
    ["T4", "P3 sale, ticket[3] = 0", "", "", "", ""],
    ["T5", "--> P2 entra a la sección crítica", "", "", "", ""],
    ["T6", "P2 sale, ticket[2] = 0", "", "", "", ""],
    ["T7", "--> P4 entra a la sección crítica", "", "", "", ""],
    ["T8", "P4 sale, ticket[4] = 0", "", "", "", ""]
]

for row_idx, row_data in enumerate(data2, 16):
    for col_idx, cell_data in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=cell_data)
        cell.border = border
        if "entra" in str(cell_data):
            cell.font = Font(color="00FF00", bold=True)
        elif "sale" in str(cell_data):
            cell.font = Font(color="FF0000", italic=True)  # Rojo para salidas

# Ajustar anchos de columna
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws.column_dimensions[col].width = 25

# Guardar el archivo
wb.save("Algoritmo_de_Lamport.xlsx")
print("Archivo Excel creado exitosamente: Algoritmo_de_Lamport.xlsx")